from __future__ import annotations

from io import BytesIO
import json
from pathlib import Path
import zipfile
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader


def _base(nome: str, tipo: str, tamanho: int) -> dict[str, Any]:
    return {
        "tipo_arquivo": tipo,
        "nome_arquivo": nome,
        "tamanho_bytes": tamanho,
        "extracao": "ESTRUTURAL",
        "publicavel_ia": False,
    }


def _interpretar_pdf(conteudo: bytes, resumo: dict[str, Any]) -> dict[str, Any]:
    leitor = PdfReader(BytesIO(conteudo))
    textos = []
    paginas_com_texto = 0
    for pagina in leitor.pages[:200]:
        texto = (pagina.extract_text() or "").strip()
        if texto:
            paginas_com_texto += 1
            textos.append(texto[:2000])
    resumo.update({
        "paginas": len(leitor.pages),
        "paginas_com_texto_extraivel": paginas_com_texto,
        "amostra_textual": "\n".join(textos)[:12000],
    })
    return resumo


def _interpretar_planilha(conteudo: bytes, resumo: dict[str, Any]) -> dict[str, Any]:
    wb = load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
    abas = []
    for nome in wb.sheetnames:
        ws = wb[nome]
        abas.append({"nome": nome, "linhas": ws.max_row, "colunas": ws.max_column})
    resumo.update({"abas": abas, "quantidade_abas": len(abas)})
    return resumo


def _texto_openxml(conteudo: bytes, prefixos: tuple[str, ...]) -> tuple[int, str]:
    total = 0
    partes: list[str] = []
    with zipfile.ZipFile(BytesIO(conteudo)) as pacote:
        for nome in pacote.namelist():
            if not nome.endswith(".xml") or not nome.startswith(prefixos):
                continue
            total += 1
            bruto = pacote.read(nome).decode("utf-8", errors="ignore")
            texto = bruto.replace("<", " <")
            import re
            texto = re.sub(r"<[^>]+>", " ", texto)
            texto = re.sub(r"\s+", " ", texto).strip()
            if texto:
                partes.append(texto[:3000])
    return total, "\n".join(partes)[:12000]


def _interpretar_word(conteudo: bytes, resumo: dict[str, Any]) -> dict[str, Any]:
    total, texto = _texto_openxml(conteudo, ("word/",))
    resumo.update({"partes_openxml": total, "amostra_textual": texto})
    return resumo


def _interpretar_powerpoint(conteudo: bytes, resumo: dict[str, Any]) -> dict[str, Any]:
    total, texto = _texto_openxml(conteudo, ("ppt/slides/",))
    resumo.update({"slides_xml": total, "amostra_textual": texto})
    return resumo


def _interpretar_texto(conteudo: bytes, resumo: dict[str, Any]) -> dict[str, Any]:
    texto = conteudo.decode("utf-8", errors="replace")
    resumo.update({"caracteres": len(texto), "amostra_textual": texto[:12000]})
    return resumo


def _interpretar_json(conteudo: bytes, resumo: dict[str, Any]) -> dict[str, Any]:
    dados = json.loads(conteudo.decode("utf-8"))
    if isinstance(dados, dict):
        resumo.update({"estrutura": "OBJETO", "chaves_raiz": list(dados.keys())[:100]})
    elif isinstance(dados, list):
        resumo.update({"estrutura": "LISTA", "itens": len(dados)})
    else:
        resumo.update({"estrutura": type(dados).__name__.upper()})
    return resumo


def interpretar_fonte(nome: str, tipo: str, conteudo: bytes) -> dict[str, Any]:
    resumo = _base(nome, tipo, len(conteudo))
    extensao = Path(nome).suffix.lower()
    try:
        if tipo == "PDF":
            return _interpretar_pdf(conteudo, resumo)
        if tipo == "PLANILHA" and extensao in {".xlsx", ".xlsm"}:
            return _interpretar_planilha(conteudo, resumo)
        if tipo == "WORD" and extensao == ".docx":
            return _interpretar_word(conteudo, resumo)
        if tipo == "POWERPOINT" and extensao == ".pptx":
            return _interpretar_powerpoint(conteudo, resumo)
        if tipo == "TEXTO":
            return _interpretar_texto(conteudo, resumo)
        if tipo == "DADOS_ESTRUTURADOS" and extensao == ".json":
            return _interpretar_json(conteudo, resumo)
        resumo["observacao"] = "Tipo reconhecido; extração estrutural profunda será realizada por adaptador compatível quando disponível."
        return resumo
    except Exception as exc:
        resumo.update({"erro_interpretacao": type(exc).__name__, "observacao": "Arquivo preservado, mas a interpretação estrutural não foi concluída."})
        return resumo
