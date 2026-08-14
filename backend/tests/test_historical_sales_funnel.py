from pathlib import Path

from openpyxl import Workbook

from parsers.historical_sales_funnel import parse_workbook, summarize


def _fixture(tmp_path: Path):
    path = tmp_path / "funil.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    headers = {
        "BACKLOG": ["REPRESENTANTE", "DATA", "RAZÃO SOCIAL", "MODELO EQUIP.", "QTD", "VALOR UNITÁRIO", "VALOR TOTAL", "MÊS DE FATURAMENTO", "%", "CASO SC", "OBSERVAÇÃO"],
        "OPORTUNIDADE": ["REPRESENTANTE", "DATA", "RAZÃO SOCIAL", "MODELO EQUIP.", "QTD", "VALOR UNITÁRIO", "VALOR TOTAL", "MÊS POSSÍVEL DE FATURAMENTO", "POSSIBILIDADE DE VITÓRIA", "OBSERVAÇÃO"],
        "INTERMEDIAÇÃO - OEM": ["REPRESENTANTE", "DATA", "RAZÃO SOCIAL", "MODELO EQUIP.", "QTD", "MÊS POSSÍVEL DE FATURAMENTO", "POSSIBILIDADE DE VITÓRIA", "OBSERVAÇÃO"],
    }
    for name, columns in headers.items():
        ws = wb.create_sheet(name)
        for col, header in enumerate(columns, 1):
            ws.cell(5, col).value = header

    backlog = wb["BACKLOG"]
    backlog.cell(6, 1).value = "CARLA - VIENA SP"
    backlog.cell(6, 2).value = "14/09/2023"
    backlog.cell(6, 3).value = "Cliente A Ltda"
    backlog.cell(6, 4).value = "Supra 850MT"
    backlog.cell(6, 5).value = 2
    backlog.cell(6, 6).value = 100
    backlog.cell(6, 7).value = 200
    backlog.cell(6, 11).value = "concluído"

    oportunidade = wb["OPORTUNIDADE"]
    oportunidade.cell(6, 1).value = "ANDERSON - viena sp"
    oportunidade.cell(6, 2).value = "20/07/2023"
    oportunidade.cell(6, 3).value = "Cliente B"
    oportunidade.cell(6, 4).value = "X4 7500"
    oportunidade.cell(6, 5).value = 1
    oportunidade.cell(6, 6).value = 300
    oportunidade.cell(6, 7).value = 300
    oportunidade.cell(6, 9).value = 0
    oportunidade.cell(6, 10).value = "PERDEU PARA A CONCORRÊNCIA"

    oem = wb["INTERMEDIAÇÃO - OEM"]
    oem.cell(6, 1).value = "CARLA"
    oem.cell(6, 2).value = "31/01/2024"
    oem.cell(6, 3).value = "Cliente C"
    oem.cell(6, 4).value = "Supra 850"
    oem.cell(6, 5).value = 10
    oem.cell(6, 6).value = 97000
    oem.cell(6, 7).value = "=E6*F6"
    oem.cell(6, 8).value = "BORTOLOTO/IBIPORÃ - FECHOU"

    wb.save(path)
    return path


def test_contract_and_succession(tmp_path):
    records = parse_workbook(_fixture(tmp_path))
    summary = summarize(records)
    assert summary["total"] == 3
    assert summary["por_aba"] == {"BACKLOG": 1, "OPORTUNIDADE": 1, "INTERMEDIAÇÃO - OEM": 1}
    assert records[0].representante_original == "CARLA - VIENA SP"
    assert records[0].representante_normalizado == "MÔNICA - VIENA SP"
    assert "REPRESENTANTE_SUBSTITUIDO_CARLA_POR_MONICA" in records[0].flags_validacao


def test_oem_is_channel_relationship_not_second_opportunity(tmp_path):
    record = [r for r in parse_workbook(_fixture(tmp_path)) if r.aba_origem == "INTERMEDIAÇÃO - OEM"][0]
    assert record.canal_venda == "INDIRETA_OEM"
    assert record.cliente_original == "Cliente C"
    assert record.implementadora_normalizada == "IBIPORÃ"
    assert record.valor_unitario == 97000
    assert record.valor_total is None
    assert "OEM_COLUNAS_FG_FINANCEIRAS" in record.flags_validacao


def test_opportunity_zero_probability_is_preserved_but_flagged(tmp_path):
    record = [r for r in parse_workbook(_fixture(tmp_path)) if r.aba_origem == "OPORTUNIDADE"][0]
    assert record.probabilidade_original == 0
    assert str(record.probabilidade_normalizada) == "0"
    assert "PROBABILIDADE_ZERO_NAO_CONFIAVEL" in record.flags_validacao
    assert record.status_normalizado == "PERDIDO"
    assert record.motivo_perda_normalizado == "CONCORRENCIA"
