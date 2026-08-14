from __future__ import annotations

from io import BytesIO
import json
from pathlib import Path
import re
from typing import Any
import zipfile

from openpyxl import load_workbook
from pypdf import PdfReader

MAX_REGISTROS = 5000
MAX_TEXTO_REGISTRO = 6000

CLASSIFICACOES = {
    "COMERCIAL": ("cliente", "oportunidade", "venda", "pedido", "proposta", "comissao", "preco", "preço", "faturamento"),
    "MERCADO_ANFIR": ("anfir", "implementadora", "emplacamento", "frota", "carroceria", "implemento"),
    "TECNICO_PRODUTO": ("equipamento", "modelo", "compressor", "refrigeracao", "refrigeração", "carrier", "transicold", "temperatura"),
    "TERRITORIAL": ("ddd", "territorio", "território", "regiao", "região", "cidade", "estado", "municipio", "município"),
    "FINANCEIRO": ("financeiro", "custo", "margem", "receita", "imposto", "valor", "pagamento"),
    "CONTRATUAL_DOCUMENTAL": ("contrato", "clausula", "cláusula", "termo", "assinatura", "proposta comercial"),
}


def _texto(valor: Any) -> str:
    return str(valor or "").strip()


def _classificar(texto: str, tipo: str) -> tuple[str, float]:
    alvo = texto.casefold()
    placar: list[tuple[int, str]] = []
    for nome, termos in CLASSIFICACOES.items():
        pontos = sum(alvo.count(termo.casefold()) for termo in termos)
        placar.append((pontos, nome))
    placar.sort(reverse=True)
    pontos, nome = placar[0]
    if pontos <= 0:
        if tipo == "IMAGEM":
            return "DOCUMENTAL_VISUAL", 0.55
        return "DOCUMENTAL_GERAL", 0.50
    confianca = min(0.98, 0.60 + pontos * 0.035)
    return nome, round(confianca, 4)


def _registro_textual(indice: int, texto: str, tipo_registro: str, metadados: dict[str, Any]) -> dict[str, Any]:
    trecho = re.sub(r"\s+", " ", texto or "").strip()[:MAX_TEXTO_REGISTRO]
    return {
        "indice": indice,
        "tipo_registro": tipo_registro,
        "conteudo_texto": trecho,
        "dados": {"texto": trecho},
        "metadados": metadados,
    }


def _chunks_texto(texto: str, *, origem: str) -> list[dict[str, Any]]:
    texto = re.sub(r"\s+", " ", texto or "").strip()
    if not texto:
        return []
    registros: list[dict[str, Any]] = []
    for indice, inicio in enumerate(range(0, len(texto), MAX_TEXTO_REGISTRO)):
        if indice >= MAX_REGISTROS:
            break
        trecho = texto[inicio:inicio + MAX_TEXTO_REGISTRO].strip()
        if trecho:
            registros.append(_registro_textual(indice, trecho, "TRECHO_TEXTUAL", {"origem": origem, "inicio_caractere": inicio}))
    return registros


def _pdf(conteudo: bytes) -> tuple[list[dict[str, Any]], list[str], str]:
    leitor = PdfReader(BytesIO(conteudo))
    saida: list[dict[str, Any]] = []
    amostras: list[str] = []
    for pagina_num, pagina in enumerate(leitor.pages, start=1):
        if len(saida) >= MAX_REGISTROS:
            break
        texto = (pagina.extract_text() or "").strip()
        if not texto:
            continue
        amostras.append(texto[:1200])
        partes = _chunks_texto(texto, origem=f"PDF_PAGINA_{pagina_num}")
        for parte in partes:
            parte["indice"] = len(saida)
            parte["tipo_registro"] = "PAGINA_PDF"
            parte["metadados"] = {"pagina": pagina_num}
            saida.append(parte)
            if len(saida) >= MAX_REGISTROS:
                break
    return saida, ["texto"], "\n".join(amostras[:80])


def _texto_openxml(conteudo: bytes, prefixo: str) -> list[tuple[str, str]]:
    saida: list[tuple[str, str]] = []
    with zipfile.ZipFile(BytesIO(conteudo)) as pacote:
        nomes = sorted(nome for nome in pacote.namelist() if nome.startswith(prefixo) and nome.endswith(".xml"))
        for nome in nomes:
            bruto = pacote.read(nome).decode("utf-8", errors="ignore")
            texto = re.sub(r"<[^>]+>", " ", bruto)
            texto = re.sub(r"\s+", " ", texto).strip()
            if texto:
                saida.append((nome, texto))
    return saida


def _openxml(conteudo: bytes, tipo: str) -> tuple[list[dict[str, Any]], list[str], str]:
    prefixo = "word/" if tipo == "WORD" else "ppt/slides/"
    partes = _texto_openxml(conteudo, prefixo)
    saida: list[dict[str, Any]] = []
    amostras: list[str] = []
    for nome_parte, texto in partes:
        amostras.append(texto[:1200])
        for trecho in _chunks_texto(texto, origem=nome_parte):
            if len(saida) >= MAX_REGISTROS:
                break
            trecho["indice"] = len(saida)
            trecho["tipo_registro"] = "TRECHO_WORD" if tipo == "WORD" else "SLIDE_POWERPOINT"
            trecho["metadados"] = {"parte_openxml": nome_parte}
            saida.append(trecho)
        if len(saida) >= MAX_REGISTROS:
            break
    return saida, ["texto"], "\n".join(amostras[:80])


def _planilha(conteudo: bytes) -> tuple[list[dict[str, Any]], list[str], str]:
    wb = load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
    saida: list[dict[str, Any]] = []
    campos: set[str] = set()
    amostras: list[str] = []
    indice = 0
    for ws in wb.worksheets:
        linhas = ws.iter_rows(values_only=True)
        cabecalho = None
        for linha_num, linha in enumerate(linhas, start=1):
            valores = list(linha)
            if not any(v not in (None, "") for v in valores):
                continue
            if cabecalho is None:
                cabecalho = [(_texto(v) or f"coluna_{i+1}")[:120] for i, v in enumerate(valores)]
                continue
            dados = {cabecalho[i]: valores[i] for i in range(min(len(cabecalho), len(valores))) if valores[i] not in (None, "")}
            if not dados:
                continue
            campos.update(dados.keys())
            amostras.append(json.dumps(dados, ensure_ascii=False, default=str)[:1000])
            saida.append({
                "indice": indice,
                "tipo_registro": "LINHA_PLANILHA",
                "conteudo_texto": " | ".join(f"{k}: {v}" for k, v in dados.items())[:MAX_TEXTO_REGISTRO],
                "dados": dados,
                "metadados": {"aba": ws.title, "linha": linha_num},
            })
            indice += 1
            if indice >= MAX_REGISTROS:
                break
        if indice >= MAX_REGISTROS:
            break
    return saida, sorted(campos), "\n".join(amostras[:80])


def _json(conteudo: bytes) -> tuple[list[dict[str, Any]], list[str], str]:
    valor = json.loads(conteudo.decode("utf-8"))
    itens = valor if isinstance(valor, list) else [valor]
    saida = []
    campos: set[str] = set()
    for indice, item in enumerate(itens[:MAX_REGISTROS]):
        dados = item if isinstance(item, dict) else {"valor": item}
        campos.update(str(k) for k in dados.keys())
        saida.append({
            "indice": indice,
            "tipo_registro": "REGISTRO_JSON",
            "conteudo_texto": json.dumps(dados, ensure_ascii=False, default=str)[:MAX_TEXTO_REGISTRO],
            "dados": dados,
            "metadados": {},
        })
    return saida, sorted(campos), json.dumps(itens[:80], ensure_ascii=False, default=str)[:60000]


def gerar_semantica(nome: str, tipo: str, conteudo: bytes, resumo_estrutural: dict[str, Any]) -> dict[str, Any]:
    extensao = Path(nome).suffix.lower()
    registros: list[dict[str, Any]] = []
    campos: list[str] = []
    texto_classificacao = _texto(resumo_estrutural.get("amostra_textual"))

    if tipo == "PDF":
        registros, campos, texto_classificacao = _pdf(conteudo)
    elif tipo == "PLANILHA" and extensao in {".xlsx", ".xlsm"}:
        registros, campos, texto_classificacao = _planilha(conteudo)
    elif tipo in {"WORD", "POWERPOINT"} and extensao in {".docx", ".pptx"}:
        registros, campos, texto_classificacao = _openxml(conteudo, tipo)
    elif tipo == "DADOS_ESTRUTURADOS" and extensao == ".json":
        registros, campos, texto_classificacao = _json(conteudo)
    elif tipo == "TEXTO":
        texto_classificacao = conteudo.decode("utf-8", errors="replace")
        registros = _chunks_texto(texto_classificacao, origem="TEXTO")
        campos = ["texto"] if registros else []
    else:
        registros = _chunks_texto(texto_classificacao, origem=tipo)
        campos = ["texto"] if registros else []

    if not registros:
        registros = [{
            "indice": 0,
            "tipo_registro": "METADADO_DOCUMENTAL",
            "conteudo_texto": f"Arquivo {nome}; tipo {tipo}; tamanho {len(conteudo)} bytes.",
            "dados": {
                "nome_arquivo": nome,
                "tipo_arquivo": tipo,
                "tamanho_bytes": len(conteudo),
                "resumo_estrutural": resumo_estrutural,
            },
            "metadados": {"sem_texto_extraivel": True},
        }]
        campos = ["nome_arquivo", "tipo_arquivo", "tamanho_bytes", "resumo_estrutural"]

    classificacao, confianca = _classificar(texto_classificacao + " " + nome, tipo)
    descricao = (
        f"Fonte dinâmica homologável '{nome}', classificada como {classificacao}. "
        f"Contém {len(registros)} registro(s) semântico(s) derivados do original preservado."
    )
    return {
        "classificacao_sugerida": classificacao,
        "confianca_classificacao": confianca,
        "descricao_semantica": descricao,
        "campos_semanticos": campos,
        "registros": registros,
        "total_registros": len(registros),
        "limite_registros_aplicado": len(registros) >= MAX_REGISTROS,
    }
