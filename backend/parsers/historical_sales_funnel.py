"""HIST-001 - parser read-only do histórico comercial 2023-2026.

Não possui integração com banco, Supabase, Pipeline, Forecast, Pedidos, Vendas ou IA.
A saída é um contrato em memória para staging futuro, preservando proveniência e valores originais.
"""
from __future__ import annotations

from dataclasses import dataclass, asdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from hashlib import sha256
import json
import re
import unicodedata
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from core.cti_taxonomy import consolidar_cliente, normalizar_implementadora, IMPLEMENTADORAS

HEADER_ROW = 5
DATA_START_ROW = 6
EXPECTED_SHEETS = ("BACKLOG", "OPORTUNIDADE", "INTERMEDIAÇÃO - OEM")
SOURCE_MAX_COLUMNS = {"BACKLOG": 15, "OPORTUNIDADE": 14, "INTERMEDIAÇÃO - OEM": 8}

SHEET_COLUMNS = {
    "BACKLOG": {
        1: "representante", 2: "data", 3: "cliente", 4: "equipamento", 5: "quantidade",
        6: "valor_unitario", 7: "valor_total", 8: "previsao", 9: "probabilidade",
        10: "caso_sc", 11: "observacao",
    },
    "OPORTUNIDADE": {
        1: "representante", 2: "data", 3: "cliente", 4: "equipamento", 5: "quantidade",
        6: "valor_unitario", 7: "valor_total", 8: "previsao", 9: "probabilidade",
        10: "observacao",
    },
    "INTERMEDIAÇÃO - OEM": {
        1: "representante", 2: "data", 3: "cliente", 4: "equipamento", 5: "quantidade",
        6: "previsao_ou_valor_unitario", 7: "probabilidade_ou_valor_total", 8: "observacao",
    },
}

REPRESENTATIVE_SUCCESSION = {"CARLA": "MÔNICA - VIENA SP"}
KNOWN_REPRESENTATIVES = {
    "ANDERSON": "ANDERSON - VIENA SP",
    "ANDRE": "ANDRE - VIENA SP",
    "NATHAN": "NATHAN - VIENA SP",
    "MICHELE": "MICHELE - VIENA SP",
    "MONICA": "MÔNICA - VIENA SP",
    "MÔNICA": "MÔNICA - VIENA SP",
}
IMPLEMENTADORA_TOKENS = tuple(sorted({*IMPLEMENTADORAS.keys(), *IMPLEMENTADORAS.values(), "BORTOLOTO", "FRATELLI", "LM", "MULTIEIXO"}, key=len, reverse=True))


def _json_safe(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _clean_display(value: Any) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def _fold(value: Any) -> str:
    text = _clean_display(value) or ""
    text = "".join(c for c in unicodedata.normalize("NFKD", text.upper()) if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", text).strip()


def normalize_representative(value: Any) -> tuple[str | None, list[str]]:
    original = _clean_display(value)
    if not original:
        return None, ["REPRESENTANTE_AUSENTE"]
    folded = _fold(original)
    if "CARLA" in folded:
        return REPRESENTATIVE_SUCCESSION["CARLA"], ["REPRESENTANTE_SUBSTITUIDO_CARLA_POR_MONICA"]
    for token, official in KNOWN_REPRESENTATIVES.items():
        if _fold(token) in folded:
            return official, []
    if folded == "VIENA SP":
        return "VIENA SP", ["REPRESENTANTE_NAO_INDIVIDUALIZADO"]
    return folded, ["REPRESENTANTE_NAO_CATALOGADO"]


def _decimal(value: Any) -> Decimal | None:
    if value is None or value == "" or isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    text = str(value).strip().replace("R$", "").replace(" ", "")
    if text.startswith("="):
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _integer(value: Any) -> int | None:
    d = _decimal(value)
    if d is None or d != d.to_integral_value():
        return None
    return int(d)


def _date(value: Any, epoch) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and 1 <= value <= 80000:
        try:
            return from_excel(value, epoch).date()
        except Exception:
            return None
    if isinstance(value, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                pass
    return None


def _normalize_equipment(value: Any) -> str | None:
    text = _clean_display(value)
    return _fold(text) if text else None


def _probability(value: Any) -> Decimal | None:
    d = _decimal(value)
    if d is None:
        return None
    if Decimal("0") <= d <= Decimal("1"):
        return d
    if Decimal("1") < d <= Decimal("100"):
        return d / Decimal("100")
    return None


def _extract_implementadora(observation: Any) -> tuple[str | None, str | None]:
    obs = _clean_display(observation)
    if not obs:
        return None, None
    folded = _fold(obs)
    hits = [token for token in IMPLEMENTADORA_TOKENS if _fold(token) in folded]
    if "BORTOLOTO" in folded and "IBIPORA" in folded:
        return "BORTOLOTO/IBIPORÃ", "IBIPORÃ"
    if not hits:
        return None, None
    original = "/".join(dict.fromkeys(hits))
    return original, normalizar_implementadora(hits[0])


def _oem_financial_semantics(raw_f: Any, raw_g: Any, cached_f: Any, cached_g: Any) -> bool:
    if isinstance(raw_g, str) and raw_g.lstrip().startswith("="):
        return True
    dg = _decimal(cached_g if cached_g is not None else raw_g)
    df = _decimal(cached_f if cached_f is not None else raw_f)
    if dg is not None and abs(dg) > Decimal("100"):
        return True
    if df is not None and abs(df) > Decimal("1000") and (raw_g is None or (dg is not None and abs(dg) > Decimal("100"))):
        return True
    return False


def _status_from_observation(sheet: str, observation: Any) -> tuple[str | None, str | None]:
    text = _fold(observation)
    if not text:
        return None, None
    if any(x in text for x in ("CANCEL", "DECLIN", "PERDEMOS", "PERDEU", "CONCORREN")):
        reason = "CONCORRENCIA" if "CONCORREN" in text else ("PRECO" if "PRECO" in text else None)
        return "PERDIDO", reason
    if "SEM RETORNO" in text:
        return "PERDIDO", "SEM_RETORNO"
    if "PRECO" in text and sheet == "OPORTUNIDADE":
        return "EM_NEGOCIACAO", "PRECO"
    if any(x in text for x in ("CONCLUID", "FINALIZ", "FECHOU", "GANHAMOS", "FECHADO")):
        return "GANHO", None
    if any(x in text for x in ("NEGOCI", "COTAC", "ANALISE")):
        return "EM_NEGOCIACAO", None
    if "SO FATURAR" in text or "FATUR" in text:
        return "FATURAMENTO_PENDENTE", None
    return None, None


@dataclass
class HistoricalRecord:
    arquivo_origem: str
    arquivo_sha256: str
    aba_origem: str
    linha_origem: int
    registro_original: dict[str, Any]
    registro_hash: str
    cliente_original: str | None
    cliente_normalizado: str | None
    representante_original: str | None
    representante_normalizado: str | None
    equipamento_original: str | None
    equipamento_normalizado: str | None
    quantidade: int | None
    valor_unitario: Decimal | None
    valor_total: Decimal | None
    data_original: Any
    data_normalizada: date | None
    previsao_original: Any
    probabilidade_original: Any
    probabilidade_normalizada: Decimal | None
    status_original: str | None
    status_normalizado: str | None
    motivo_perda_normalizado: str | None
    observacao_original: str | None
    canal_venda: str
    implementadora_original: str | None
    implementadora_normalizada: str | None
    flags_validacao: list[str]

    def as_dict(self) -> dict[str, Any]:
        result = asdict(self)
        result["registro_original"] = {k: _json_safe(v) for k, v in self.registro_original.items()}
        for key, value in list(result.items()):
            if key != "registro_original":
                result[key] = _json_safe(value)
        return result


def file_sha256(path: str | Path) -> str:
    h = sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _row_original(ws, row: int, sheet: str) -> dict[str, Any]:
    result = {}
    for col in range(1, SOURCE_MAX_COLUMNS[sheet] + 1):
        header = ws.cell(HEADER_ROW, col).value
        key = f"{ws.cell(HEADER_ROW, col).column_letter}:{_clean_display(header) or 'SEM_CABECALHO'}"
        result[key] = ws.cell(row, col).value
    return result


def _is_business_row(ws, row: int) -> bool:
    return any(ws.cell(row, c).value not in (None, "") for c in range(1, min(ws.max_column, 8) + 1)) and ws.cell(row, 3).value not in (None, "")


def parse_workbook(path: str | Path) -> list[HistoricalRecord]:
    path = Path(path)
    file_hash = file_sha256(path)
    wb_raw = load_workbook(path, data_only=False, read_only=False)
    wb_cached = load_workbook(path, data_only=True, read_only=False)
    missing = [s for s in EXPECTED_SHEETS if s not in wb_raw.sheetnames]
    if missing:
        raise ValueError(f"Abas obrigatórias ausentes: {missing}")
    records = []
    for sheet in EXPECTED_SHEETS:
        ws = wb_raw[sheet]
        wc = wb_cached[sheet]
        for row in range(DATA_START_ROW, ws.max_row + 1):
            if not _is_business_row(ws, row):
                continue
            raw = _row_original(ws, row, sheet)
            rep_o = _clean_display(ws.cell(row, 1).value)
            rep_n, flags = normalize_representative(rep_o)
            client_o = _clean_display(ws.cell(row, 3).value)
            equip_o = _clean_display(ws.cell(row, 4).value)
            qty = _integer(wc.cell(row, 5).value)
            obs_col = 11 if sheet == "BACKLOG" else 10 if sheet == "OPORTUNIDADE" else 8
            obs = _clean_display(ws.cell(row, obs_col).value)
            value_u = value_t = forecast_o = prob_o = None
            if sheet != "INTERMEDIAÇÃO - OEM":
                value_u = _decimal(wc.cell(row, 6).value)
                value_t = _decimal(wc.cell(row, 7).value)
                forecast_o = ws.cell(row, 8).value
                prob_o = wc.cell(row, 9).value
            else:
                rf, rg = ws.cell(row, 6).value, ws.cell(row, 7).value
                cf, cg = wc.cell(row, 6).value, wc.cell(row, 7).value
                if _oem_financial_semantics(rf, rg, cf, cg):
                    value_u = _decimal(cf if cf is not None else rf)
                    value_t = _decimal(cg if cg is not None else rg)
                    if value_t is None and value_u is not None and qty is not None:
                        flags.append("VALOR_TOTAL_AUSENTE_NAO_INFERIDO")
                    flags.append("OEM_COLUNAS_FG_FINANCEIRAS")
                else:
                    forecast_o = rf
                    prob_o = cg if cg is not None else rg
            if value_u is not None and value_t is not None and qty is not None and value_u * qty != value_t:
                flags.append("DIVERGENCIA_ARITMETICA")
            data_o = ws.cell(row, 2).value
            data_n = _date(wc.cell(row, 2).value, wb_cached.epoch)
            if data_n is None:
                flags.append("DATA_INVALIDA")
            prob_n = _probability(prob_o)
            if sheet == "OPORTUNIDADE" and prob_n == Decimal("0"):
                flags.append("PROBABILIDADE_ZERO_NAO_CONFIAVEL")
            status_n, reason = _status_from_observation(sheet, obs)
            impl_o = impl_n = None
            channel = "INDIRETA_OEM" if sheet == "INTERMEDIAÇÃO - OEM" else "DIRETA"
            if channel == "INDIRETA_OEM":
                impl_o, impl_n = _extract_implementadora(obs)
                if not impl_o:
                    flags.append("IMPLEMENTADORA_NAO_IDENTIFICADA")
            payload = {"arquivo": path.name, "aba": sheet, "linha": row, "registro_original": {k: _json_safe(v) for k, v in raw.items()}}
            row_hash = sha256(json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str).encode()).hexdigest()
            records.append(HistoricalRecord(
                arquivo_origem=path.name, arquivo_sha256=file_hash, aba_origem=sheet, linha_origem=row,
                registro_original=raw, registro_hash=row_hash,
                cliente_original=client_o, cliente_normalizado=consolidar_cliente(client_o) if client_o else None,
                representante_original=rep_o, representante_normalizado=rep_n,
                equipamento_original=equip_o, equipamento_normalizado=_normalize_equipment(equip_o),
                quantidade=qty, valor_unitario=value_u, valor_total=value_t,
                data_original=data_o, data_normalizada=data_n, previsao_original=forecast_o,
                probabilidade_original=prob_o, probabilidade_normalizada=prob_n,
                status_original=obs, status_normalizado=status_n, motivo_perda_normalizado=reason,
                observacao_original=obs, canal_venda=channel,
                implementadora_original=impl_o, implementadora_normalizada=impl_n,
                flags_validacao=flags,
            ))
    return records


def summarize(records: Iterable[HistoricalRecord]) -> dict[str, Any]:
    rows = list(records)
    by_sheet = {s: 0 for s in EXPECTED_SHEETS}
    for record in rows:
        by_sheet[record.aba_origem] += 1
    return {
        "total": len(rows),
        "por_aba": by_sheet,
        "direta": sum(r.canal_venda == "DIRETA" for r in rows),
        "indireta_oem": sum(r.canal_venda == "INDIRETA_OEM" for r in rows),
        "carla_reconciliada_para_monica": sum("REPRESENTANTE_SUBSTITUIDO_CARLA_POR_MONICA" in r.flags_validacao for r in rows),
        "divergencias_aritmeticas": sum("DIVERGENCIA_ARITMETICA" in r.flags_validacao for r in rows),
        "oem_fg_financeiras": sum("OEM_COLUNAS_FG_FINANCEIRAS" in r.flags_validacao for r in rows),
    }
