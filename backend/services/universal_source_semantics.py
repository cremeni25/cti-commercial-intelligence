from __future__ import annotations

from io import BytesIO
import json
from pathlib import Path
import re
import zipfile
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader


CLASSIFICACOES = {
    "COMERCIAL_HISTORICO": ("oportunidade", "backlog", "venda", "pedido", "cliente", "representante"),
    "ANFIR_MERCADO": ("anfir", "implementadora", "carroceria", "emplacamento", "frota"),
    "PROPOSTA_COMERCIAL": ("proposta", "condições comerciais", "validade da proposta", "preço unitário"),
    "PEDIDO_VENDA": ("pedido", "ordem de compra", "faturamento", "entrega"),
    "RELATORIO_COMERCIAL": ("relatório", "relatorio", "forecast", "pipeline", "meta", "resultado"),
    "CADASTRO_EMPRESA": ("cnpj", "razão social", "razao social", "endereço", "endereco"),
    "MATERIAL_TECNICO": ("especificação", "especificacao", "compressor", "capacidade", "refrigeração", "refrigeracao"),
}


def _texto_limpo(valor: Any) -> str:
    return re.sub(r"\s+", " ", str(valor or "")).strip()


def _classificar(texto: str, nome: str) -> tuple[str, float, list[str]]:
    alvo = f"{nome} {texto}".casefold()
    pontuacoes: list[tuple[int, str, list[str]]] = []
    for classificacao, termos in CLASSIFICACOES.items():
        encontrados = [termo for termo in termos if termo.casefold() in alvo]
        pontuacoes.append((len(encontrados), classificacao, encontrados))
    pontuacoes.sort(reverse=True)
    pontos, classe, encontrados = pontuacoes[0]
    if pontos == 0:
        return "DOCUMENTO_GERAL", 0.45, []
    confianca = min(0.98, 0.58 + pontos * 0.08)
    return classe, round(confianca, 4), encontrados


def _registro(indice: int, tipo: str, texto: str, dados: dict[str, Any], metadados: dict[str, Any]) -> dict[str, Any]:
    return {
        "indice": indice,
        "tipo_registro": tipo,
        "conteudo_texto": texto[:12000],
        "dados": dados,
        "metadados": metadados,
    }


def _pdf(conteudo: bytes) -> list[dict[str, Any]]:
    leitor = PdfReader(BytesIO(conteudo))
    saida = []
    for i, pagina in enumerate(leitor.pages[:500], start=1):
        texto = _texto_limpo(pagina.extract_text() or "")
        saida.append(_registro(i, "PAGINA", texto, {"pagina": i}, {}))
    return saida


def _xlsx(conteudo: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
    saida: list[dict[str, Any]] = []
    indice = 0
    for nome in wb.sheetnames[:100]:
        ws = wb[nome]
        rows = ws.iter_rows(values_only=True)
        try:
            cabecalho_bruto = next(rows)
        except StopIteration:
            continue
        cabecalho = [_texto_limpo(v) or f"coluna_{i+1}" for i, v in enumerate(cabecalho_bruto)]
        for numero_linha, valores in enumerate(rows, start=2):
            if numero_linha > 10002:
                break
            if not any(v not in (None, "") for v in valores):
                continue
            indice += 1
            dados = {cabecalho[i]: valores[i] if i < len(valores) else None for i in range(len(cabecalho))}
            texto = " | ".join(f"{k}: {_texto_limpo(v)}" for k, v in dados.items() if v not in (None, ""))
            saida.append(_registro(indice, "LINHA_PLANILHA", texto, dados, {"aba": nome, "linha": numero_linha}))
    return saida


def _openxml_textos(conteudo: bytes, prefixo: str, tipo: str) -> list[dict[str, Any]]:
    saida = []
    with zipfile.ZipFile(BytesIO(conteudo)) as pacote:
        nomes = sorted(n for n in pacote.namelist() if n.startswith(prefixo) and n.endswith(".xml"))
        for i, nome in enumerate(nomes, start=1):
            bruto = pacote.read(nome).decode("utf-8", errors="ignore")
            texto = re.sub(r"<[^>]+>", " ", bruto)
            texto = _texto_limpo(texto)
            saida.append(_registro(i, tipo, texto, {}, {"parte": nome}))
    return saida


def _texto(conteudo: bytes) -> list[dict[str, Any]]:
    bruto = conteudo.decode("utf-8", errors="replace")
    blocos = [b.strip() for b in re.split(r"\n\s*\n", bruto) if b.strip()]
    if not blocos:
        blocos = [bruto]
    return [_registro(i, "BLOCO_TEXTO", _texto_limpo(bloco), {}, {}) for i, bloco in enumerate(blocos[:5000], start=1)]


def _json(conteudo: bytes) -> list[dict[str, Any]]:
    dados = json.loads(conteudo.decode("utf-8"))
    itens = dados if isinstance(dados, list) else [dados]
    saida = []
    for i, item in enumerate(itens[:10000], start=1):
        obj = item if isinstance(item, dict) else {"valor": item}
        texto = json.dumps(obj, ensure_ascii=False, default=str)
        saida.append(_registro(i, "REGISTRO_JSON", texto, obj, {}))
    return saida


def interpretar_semanticamente(nome: str, tipo: str, conteudo: bytes) -> dict[str, Any]:
    ext = Path(nome).suffix.lower()
    if tipo == "PDF":
        registros = _pdf(conteudo)
    elif tipo == "PLANILHA" and ext in {".xlsx", ".xlsm"}:
        registros = _xlsx(conteudo)
    elif tipo == "WORD" and ext == ".docx":
        registros = _openxml_textos(conteudo, "word/", "PARTE_WORD")
    elif tipo == "POWERPOINT" and ext == ".pptx":
        registros = _openxml_textos(conteudo, "ppt/slides/", "SLIDE")
    elif tipo == "TEXTO":
        registros = _texto(conteudo)
    elif tipo == "DADOS_ESTRUTURADOS" and ext == ".json":
        registros = _json(conteudo)
    else:
        registros = [_registro(1, "ARQUIVO", "", {}, {"tipo": tipo, "extensao": ext})]

    amostra = "\n".join(r["conteudo_texto"] for r in registros[:60] if r.get("conteudo_texto"))[:30000]
    classificacao, confianca, termos = _classificar(amostra, nome)
    campos: set[str] = set()
    for r in registros[:500]:
        dados = r.get("dados") or {}
        if isinstance(dados, dict):
            campos.update(str(k) for k in dados.keys())
    descricao = (
        f"Fonte {classificacao.lower().replace('_', ' ')} derivada de {nome}. "
        f"Contém {len(registros)} registro(s) semântico(s) consultáveis após homologação e publicação."
    )
    return {
        "classificacao_sugerida": classificacao,
        "confianca_classificacao": confianca,
        "termos_classificacao": termos,
        "descricao_semantica": descricao,
        "campos_semanticos": sorted(campos),
        "total_registros_semanticos": len(registros),
        "preview": registros[:30],
        "registros": registros,
    }
